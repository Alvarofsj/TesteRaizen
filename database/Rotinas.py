#!/usr/bin/env python
#*- coding: utf-8 -*-
from Banco       import *
from Postos      import *
from Postos_SMAP import *
#-------------------------------------------
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func
from zipfile import *
from datetime import *
import requests
import pandas as pd
import openpyxl as pyxl
import pyexcel
import shutil
import math

class Hidrologia(object):
    def __init__(self, config):
        self.config = config

    def get_rdh(self, rdhs: list, coords: dict={'row': [8, 172], 'col': [1, 28]}):

        aux = []
        #  Itera sobre rdhs passados
        for k, rdh in enumerate(rdhs):

            # Abre arquivo
            #wb = pyxl.load_workbook(
            #    filename=rdh,
            #    read_only=True,
            #)
            
            try:
                wb = pyxl.load_workbook(
                    filename=rdh,
                    read_only=True,
                )
            except:
                
                pyexcel.save_book_as(
                    file_name=rdh,
                    dest_file_name='{}.xlsx'.format(rdh[:-4])
                )
            
                wb = pyxl.load_workbook(
                    filename='{}.xlsx'.format(rdh[:-4]),
                    read_only=True,
                )
            
            ws = wb.worksheets[0]
            dt = pd.to_datetime(ws.cell(row=2, column=21).value[-10:], format='%d/%m/%Y')

            for i, row in enumerate(
                    ws.iter_rows(
                        min_row=coords['row'][0], max_row=coords['row'][1],
                        min_col=coords['col'][0], max_col=coords['col'][1]
                    )
            ):
                # Testes de verificacao
                if row[4].data_type == 'n':

                    try:  # vazao natural
                        val_vaz_natr = float(row[13].value)
                    except(ValueError, TypeError):
                        val_vaz_natr = 0.0

                    try:  # cota
                        val_cota = float(row[14].value)
                    except(ValueError, TypeError):
                        val_cota = 0.0

                    try:  # arm
                        val_arm = float(row[15].value) / 100
                    except(ValueError, TypeError):
                        val_arm = 0.0

                    try:  # vazao vertida
                        val_vaz_vert = float(row[18].value)
                    except(ValueError, TypeError):
                        val_vaz_vert = 0.0

                    try:  # vazao defluida
                        val_vaz_defl = float(row[20].value)
                    except(ValueError, TypeError):
                        val_vaz_defl = 0.0

                    try:  # vazao afluente
                        val_vaz_aflu = float(row[22].value)
                    except(ValueError, TypeError):
                        val_vaz_aflu = 0.0

                    try:  # vazao incremental
                        val_vaz_incr = float(row[23].value)
                    except(ValueError, TypeError):
                        val_vaz_incr = 0.0

                    aux.append(
                        dict(
                            num_posto=row[4].value,
                            dat_medicao=dt,
                            val_vaz_natr=val_vaz_natr,
                            val_cota=val_cota,
                            val_arm=val_arm,
                            val_vaz_vert=val_vaz_vert,
                            val_vaz_defl=val_vaz_defl,
                            val_vaz_aflu=val_vaz_aflu,
                            val_vaz_incr=val_vaz_incr
                        )
                    )

            wb.close()


        # Remocao das duplicatas
        self.dados = pd.DataFrame(aux)
        self.origem = 1

        print('RDHs lidos')
        
        return

    def get_acomph(self, acomphs, config_acomph: dict={'row': [6, 35], 'bloco_dados': 8}):
        df_vazao = pd.DataFrame()

        # Itera sobre todos os acomphs passados
        for acomph in acomphs:

            try:
                wb = pyxl.load_workbook(
                    filename=acomph,
                    read_only=False,
                    data_only=True
                )
            except:
                pyexcel.save_book_as(
                    file_name=acomph,
                    dest_file_name='{}.xlsx'.format(acomph[:-4])
                )

                wb = pyxl.load_workbook(
                    filename='{}.xlsx'.format(acomph[:-4]),
                    read_only=False,
                    data_only=True
                )

            # Itera sobre todas as worksheets
            for sheet in wb.worksheets:
                dat_medicao = []
                
                for row in sheet.iter_cols(
                    min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                    min_col=1, max_col=1
                ):
                    [dat_medicao.append(cell.value) for cell in row]
                
                # pega dados do posto
                for i in list(range(9, sheet.max_column - 1 + 9, config_acomph['bloco_dados'])):
                    vaz_natr = []
                    vaz_incr = []
                    vaz_aflu = []
                    vaz_defl = []
                    val_cota = []
                    num_posto = []
                    
                    # vazao natural
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i, max_col=i
                    ):
                        [vaz_natr.append(float(cell.value)) for cell in row]

                    # vazao incremental
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-1, max_col=i-1
                    ):
                        [vaz_incr.append(float(cell.value)) for cell in row]

                    # vazao afluente
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-2, max_col=i-2
                    ):
                        [vaz_aflu.append(float(cell.value)) for cell in row]

                    # vazao defluente
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-4, max_col=i-4
                    ):
                        [vaz_defl.append(float(cell.value)) for cell in row]

                    # vazao cota
                    for row in sheet.iter_cols(
                            min_row=config_acomph['row'][0], max_row=config_acomph['row'][1],
                            min_col=i-6, max_col=i-6
                    ):
                        [val_cota.append(float(cell.value)) for cell in row]

                    # numero do posto
                    [num_posto.append(int(sheet.cell(row=1, column=i).value)) for j in range(val_cota.__len__())]

                    df_aux = pd.DataFrame.from_dict(
                        dict(
                            num_posto=num_posto,
                            dat_medicao=dat_medicao,
                            val_vaz_natr=vaz_natr,
                            val_vaz_incr=vaz_incr,
                            val_vaz_aflu=vaz_aflu,
                            val_vaz_defl=vaz_defl,
                            val_cota=val_cota
                        )
                    )

                    if num_posto[0] != None:
                        df_vazao = pd.concat([df_vazao, df_aux])

                    if num_posto[0] == 66:
                        print('aiu')
                
            os.remove('{}.xlsx'.format(acomph[:-4]))

        #df_vazao.to_csv(
        #    path_or_buf='acomph.csv', sep=';',
        #    decimal=',', index=False,
        #    float_format='%5.2f',
        #    date_format='%Y-%m-%d'
        #)
        self.dados = pd.DataFrame(df_vazao)
        self.origem = 2

        return
        
    def get_ipdo(self, ipdos):
        
        # Colocando dados em um unico vetor
        vIpdo = []
        
        for k, ipdo in enumerate(ipdos):
    
            try:
                wb = pyxl.load_workbook(
                    filename=ipdo,
                    read_only=True
                )
            except:
                
                pyexcel.save_book_as(
                    file_name=ipdo,
                    dest_file_name='{}.xlsx'.format(ipdo[:-4])
                )
            
                wb = pyxl.load_workbook(
                    filename='{}.xlsx'.format(ipdo[:-4]),
                    read_only=True,
                )
            
            ws = wb.worksheets[1]
            dt = pd.to_datetime(ws.cell(row=6, column=20).value, format='%d/%m/%Y').to_pydatetime() # Data do relatorio
            
            ena   = [ws.cell(65, 13).value,ws.cell(64, 13).value,ws.cell(63, 13).value,ws.cell(62, 13).value] # SE,S,NE,N
            arm   = [ws.cell(65, 18).value,ws.cell(64, 18).value,ws.cell(63, 18).value,ws.cell(62, 18).value] # SE,S,NE,N
            carga = [ws.cell(40, 15).value,ws.cell(48, 15).value,ws.cell(32, 15).value,ws.cell(24, 15).value] # SE,S,NE,N
            sub   = ["SE","S","NE","N"]
            
            for i,n in enumerate(ena,0):
                
                vIpdo.append(
                            dict(
                                dat_medicao=pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                                num_ssis=int(i+1),
                                nom_ssis=sub[i],
                                ena_ssis=n,
#                               arm_ssis=arm[i].replace(",","."),
                                arm_ssis=arm[i] * 100,
                                carga_ssis=carga[i]
                                )
                )
                
            #wb.close()
        
        self.df_ipdo = pd.DataFrame(vIpdo)
        
        # Armazenamento individual ------------------------------
        
        vIpdo = list()
        
        for i in range(600,1300):
                
            vmin = i
            if ws.cell(i,13).value == "Corumbá IV":
                vmax = vmin + 42
                break
            
        # Primeira tabela
        for i in range(vmin, vmax):
            
            vIpdo.append(dict(
                                dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                                origem      = ws.cell(i,13).value,
                                val_arm     = ws.cell(i,20).value,
                              )
                         )
            
        vmin = vmax + 6
        vmax = vmin + 44
            
        # Segunda tabela
        for i in range(vmin, vmax):
        
            vIpdo.append(dict(
                                dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                                origem      = ws.cell(i,13).value,
                                val_arm     = ws.cell(i,20).value,
                              )
                         )
        
        # Dados do Grande
        vIpdo.append(dict(
                           dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                           origem      = "Grande",
                           val_arm     = ws.cell(vmax+2,13).value,
                         )
                    )
                    
        # Dados do Paranaiba
        vIpdo.append(dict(
                           dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                           origem      = "Paranaiba",
                           val_arm     = ws.cell(vmax+1,13).value,
                         )
                    )
        
        # Dados Tiete
        vIpdo.append(dict(
                           dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                           origem      = "Tiete",
                           val_arm     = ws.cell(vmax+3,13).value,
                         )
                    )
                    
        # Dados Paranapanema
        vIpdo.append(dict(
                           dat_medicao = pd.to_datetime(dt,format="%Y-%m-%d").to_pydatetime(),
                           origem      = "Paranapanema",
                           val_arm     = ws.cell(vmax+4,13).value,
                         )
                    )
                    
        wb.close()
        
            
        df_arm_ipdo = pd.DataFrame(vIpdo)
        
        self.df_arm_ipdo = df_arm_ipdo
        
        return
    
    def get_pvv(self,pvvs):
        '''
        Descompacta e le dados de historico de vazao semanal do Previvaz
        '''
        
        tmpPath = config.paths['tmp'] # Diretorio de arquivos temporarios
        
        vDados = list()
        for i,n in enumerate(pvvs,0):
            
            date_file = datetime.fromtimestamp(os.stat(n).st_mtime)
            
            zf  = ZipFile(n,"r")                 # Arquivo original do previvaz
            zf.extract(zf.namelist()[1],tmpPath) # Extraindo apenas a pasta com os arquivos de entrada
            zf.close()                           # Fechando o arquivo original
            
            folder = [x for x in os.listdir(tmpPath) if os.path.isdir(os.path.join(tmpPath,x)) == True]
            
            nzipPath = os.path.join(tmpPath,folder[0],'Arq_Entrada.zip')
            
            zf  = ZipFile(nzipPath,"r") # Arquivo original do previvaz
            zf.extractall(tmpPath)      # Extraindo apenas a pasta com os arquivos de entrada
            zf.close()                  # Fechando o arquivo original
            
            histPath = os.path.join(tmpPath,"Arq_Entrada")                                                  # Caminho onde encontram-se os arquivos de historico
            files    = [os.path.join(histPath,x) for x in os.listdir(histPath) if x[-3:] in ['DAT','dat']]  # Lista arquivos de historico
            
            for f in files:
                
                file  = open(f,"r")
                dados = file.readlines()
                file.close()
                
                if f.find("168") != -1: # Arquivo da UHE Sobradinho contem um padrao diferente
                    num_posto = int(dados[0][0:10].strip()) # Numero do posto
                    nom_posto = dados[0][10:].strip()       # Nome do posto
                else:
                    num_posto = int(dados[0][0:9].strip())  # Numero do posto
                    nom_posto = dados[0][9:].strip()        # Nome do posto
                
                ano_ini   = int(dados[1][0:5].strip())        # Ano inicio do historico
                ano_fim   = int(dados[1][6:10].strip())       # Ano fim do historico
                area_dren = int(float(dados[1][10:].strip())) # Area de drenagem do posto
                
                del dados[0], dados[0] # Exclui as duas primeiras linhas que nao contem dados de historico
                
                ano = ano_ini # Inicia variavel do ano
                
                num_semana = 1
                
                for d in dados:
                    
                    aux = d.strip().split()     # Quebra a linha em vetor
                    
                    if d[72:].strip() != "":
                        ano = int(d[72:].strip()) # Ano de medicao
                        aux = aux[:-1]            # Retira o ano do vetor de vazoes
                    
                    aux = [int(float(x)) for x in aux] # Transforma dados de vazao em INT
                    
                    for v in aux:
                        
                        vDados.append(dict(
                                            dat_medicao = date_file.date(),
                                            num_posto   = num_posto,
                                            num_semana  = num_semana,
                                            num_ano     = ano,
                                            val_vaz     = v
                                            )
                                        )
                        
                        if num_semana == config.control['num_semana']:
                            num_semana = 1
                        else:
                            num_semana += 1
                    
        df_hist      = pd.DataFrame(vDados)
        self.df_hist = df_hist
        #df_hist.to_csv("historico_vazoes_semanais.csv",sep=";",decimal=",",index=False)
        
        # Apaga dados da pasta temporaria -------
        for f in [x for x in os.listdir(tmpPath) if os.path.isdir(os.path.join(tmpPath,x)) == True]:
            
            shutil.rmtree(os.path.join(tmpPath,f))
        #----------------------------------------
        
        return
    
    def get_pmedia(self, pmedias,modelo):
        
        tmpPath = config.paths['tmp'] # Diretorio de arquivos temporarios
        
        vChuva=[] # Inicializando vetor para dados de chuva
        
        df_pm = pd.read_csv(config.paths['grade_pmedia'],sep=";",dtype={'lat': float, 'lon': float})
        
        if modelo =="eta+gefs":
            modelo3 = "pmedia"
        else:
            modelo3 = modelo
        
        df_gr = pd.read_csv(config.paths['grade_'+modelo3],sep=";",dtype={'lat': float, 'lon': float})
        df_gr['pt'] = df_gr['lon'].astype(str) + df_gr['lat'].astype(str) # Coluna de check
        
        max_lat = df_gr['lat'].astype(float).max() # Maxima latitude
        min_lat = df_gr['lat'].astype(float).min() # Minima latitude
        max_lon = df_gr['lon'].astype(float).max() # Maxima longitude
        min_lon = df_gr['lon'].astype(float).min() # Minima longitude
        
        for i,n in enumerate(pmedias,0):
            
            zf = ZipFile(n,"r")    # Declarando arquivo de chuva .zip
            zf.extractall(tmpPath) # Extraindo arquivo tempariamente para diretorio temporario
            
            #----------------------
            if modelo == "eta+gefs":
                
                modelo2 = "PMEDIA_p"
            
            elif modelo == "gefs":
                
                modelo2 = "GEFS_p"
            
            else:
                
                modelo2 = "ECMWF_p"
            #----------------------
            
            if modelo == "gefs" and os.path.exists(os.path.join(tmpPath,"datsum")): # Arquivo do GEFS contem pasta "datsum"
                files = [os.path.join("datsum",x) for x in os.listdir(os.path.join(tmpPath,"datsum"))]
            else:
                files = zf.namelist()
            
            
            for m in files: # Para cada um dos arquivos no arquivo compactado
                
                fPath = os.path.join(tmpPath,m) # Arquivo m extraido
                
                if m.find(modelo2) != -1:
                    
                    file = open(fPath,"r")          # Abre o arquivo m
                    dados = file.readlines()        # Le dados do arquivo m e coloca-os no vetor dados
                    file.close()                    # Fecha arquivo m
                    
                    pchar = m.find("p")
                    achar = m.find("a")
                    
                    datai = datetime.strptime(m[pchar+1:pchar+7],"%d%m%y").date() # Data do arquivo de previsao
                    dataf = datetime.strptime(m[pchar+8:pchar+14],"%d%m%y").date() # Data de previsao do arquivo de previsao
                    
                    for j,p in enumerate(dados,0): # Lendo todos os dados do arquivo m
                        
                        aux = p.split()                      # Separa os dados da linha por campo LON, LAT e PREC
                        aux = [float(x) for x in aux]        # Transforma numeros em float
                        aux = [datai,dataf]+aux+[modelo]     # Une o vetor em um so

                        if float(aux[2]) <= max_lon and float(aux[2]) >= min_lon \
                        and float(aux[3]) <= max_lat and float(aux[3]) >= min_lat: # Ponto de grade no limite
                            
                            vChuva.append(aux) # Adiciona no vetor auxiliar as datas, mais as informacoes do arquivo
                
                try:
                    try:
                        os.remove(fPath) # Remove arquivo m do diretorio temporario
                    except:
                        shutil.rmtree(fPath) # Remove diretorio m do diretorio temporario
                except:
                    print("Arquivo "+fPath+" nao esta mais disponivel.")
            
            # Exclui diretorios 'lixo' -------------------------------------
            folders = [os.path.join(tmpPath,x) for x in os.listdir(tmpPath) if os.path.isdir(os.path.join(tmpPath,x))]
            
            for f in folders:
                try:
                    shutil.rmtree(f)
                except:
                    print(f"Pasta {f} ja removida")
            #---------------------------------------------------------------
            
            
        zf.close()
        
        df_dad  = pd.DataFrame(vChuva,columns=["data_gera","data_prev","lon","lat","prec","modelo"]) # Dataframe com dados de chuva do arquivo f
        df_dad['pt'] = df_dad['lon'].astype(str) + df_dad['lat'].astype(str)                         # Insere coluna de check
        #df_dad.to_csv("dados_grade.csv",sep=";",decimal=".",index=False)
        df_ndad = df_dad.loc[df_dad['pt'].isin(df_gr['pt'].astype(str).tolist())]
        df_ndad = df_ndad.reset_index().drop("index",1)
        
        df_fim = pd.merge(df_gr,df_ndad,on="pt",how="left")
        #df_fim.to_csv(os.path.join(config.paths['app'],"fim-"+str(datai)+".csv"),sep=";",index=False)
        vChuva =[]
        
        for m,d in enumerate(df_pm['nom_subbac'],0):
            
            df_aux       = df_fim[df_fim['nom_subbac']==d].reset_index().drop("index",1)
            df_data_gera = df_aux[['data_gera','data_prev']].drop_duplicates().reset_index().drop("index",1)
            
            for n,e in enumerate(df_data_gera['data_gera'],0):
                
                #df_data_prev = df_aux[['data_gera','data_prev']].drop_duplicates().reset_index().drop("index",1)
                
                #for o,f in enumerate(df_data_gera['data_prev'],0):
                    
                df_aux_2 = df_aux[(df_aux['data_gera'] == e) & (df_aux['data_prev'] == df_data_gera['data_prev'][n])].reset_index().drop("index",1)
                med_prec = round(df_aux_2['prec'].astype(float).mean(),2)
                
                check = float(med_prec)
                
                if not math.isnan(check):
                    
                    aux = [e,df_data_gera['data_prev'][n],df_pm['lon'][m],df_pm['lat'][m],med_prec,modelo]
                    vChuva.append(aux)
            
        #----------------------------------------------
        
        self.df_ch = pd.DataFrame(vChuva,columns=["data_gera","data_prev","lon","lat","prec","modelo"])           # Transforma vetor em dataframe
        #self.df_ch.to_csv(os.path.join(config.paths['app'],"chuva-"+str(datai)+".csv"),sep=";",index=False) # Passa dataframe para CSV
        
        return
    
    def download_merge(self):
        
        credentials = dict(
            username=config.credentials['user_thunder'], # Usuario
            password=config.credentials['pwd_thunder']   # Senha
        )
        
        now = datetime.now().date()   # Data de hoje
        ano = str(now.year)           # Ano de hoje
        mes = str(now.month).zfill(2) # Mes de hoje
        dia = str(now.day).zfill(2)   # Mes de hoje
        
        df_hist = pd.read_csv(config.paths['merge'],sep=";") # Arquivo historico
        filename = "chuva-"+ano+mes                          # Nome do arquivo
        
        if df_hist[df_hist['path']==os.path.join(config.paths['paths_merge'],filename+dia+'.csv')].empty: # Se o arquivo ainda nao existir
            
            url = 'https://thundermeteo.com/dados/merge/'+filename+'.txt'  # URL de acesso ao arquivo
            
            r = requests.get(url, auth=(credentials['username'],credentials['password'])) # Request the file
            
            #-----------------------------------------------
            if r.status_code == 200: # Se arquivo existir
                
                with open(os.path.join(config.paths['paths_merge'],filename+dia+'.csv'), 'wb') as out:
                
                    for bits in r.iter_content():
                    
                        out.write(bits)
            #-----------------------------------------------
    def get_merge(self,merges):
        
        df_hist = pd.read_csv(config.paths['merge'],sep=";") # Arquivo historico
        
        for i,n in enumerate(merges,0):
            
            now = datetime.strptime(n[-12:-4],"%Y%m%d").date() # Data de hoje
            ano = str(now.year)                                # Ano de hoje
            mes = str(now.month).zfill(2)                      # Mes de hoje
            dia = str(now.day).zfill(2)                        # Mes de hoje
            
            df = pd.read_csv(n,sep=";") # Le o arquivo CVS baixado

            self.df_merge = pd.DataFrame([]) # Inicializando dataframe
            
            if datetime.strptime(str(df['Dia     '][len(df)-1]),"%Y%m%d").date() == now: 
                ''' Realiza procedimento apenas se ja houver publicacao dos valores do dia corrente '''
                
                total = []                                # Inicializando vetor auxiliar
                
                # Removendo caracteres especiais --------------------------------
                df.columns = [x.replace("Ê","E") for x in df.columns.values]
                df.columns = [x.replace("Í","I") for x in df.columns.values]
                df.columns = [x.replace("Ã","A") for x in df.columns.values]
                df.columns = [x.replace("Ú","U") for x in df.columns.values]
                df.columns = [x.replace("Á","A") for x in df.columns.values]
                df.columns = [x.replace("Ç","C") for x in df.columns.values]
                #----------------------------------------------------------------
                
                for i,n in enumerate(df.columns.values[1:],0): # Para cada coluna do dataframe
                    
                    for j,m in enumerate(df[n],0): # Para cada linha do dataframe
                        
                        data = datetime.strptime(str(df['Dia     '][j]),"%Y%m%d") # Passando data para formato datetime
                        
                        total.append([n.strip(),data,float(df[n][j])]) # Nome da bacia, data, precipitacao observada
                
                self.df_merge = pd.DataFrame(total,columns=["nom_bacia","data_obs","prec"]) # Transformando vetor auxiliar em dataframe para inserir no BD
                self.df_merge.to_csv(os.path.join(config.paths['paths_merge'],"merge-"+str(ano)+str(mes)+str(dia)+".csv"),sep=";",index=True)  # Transformando o dataframe em CSV
            
            else:
                
                df_hist = df_hist.drop(len(df_hist)-1)
                df_hist.to_csv(config.paths['merge'],sep=";",index=False)
        
        return
        
    def get_smap(self,smaps):
        
        vSmap = []
        for smap in smaps:
            
            df_smap = pd.read_csv(smap,sep=";",decimal=",")
            
            data_arq = smap[-12:-4]
            data_arq = datetime.strptime(data_arq,"%Y%m%d")
            
            def escreve_arquivo(dado):
                
                vSmap.append(dict(
                                dat_medicao  = data_arq,
                                nom_bacia    = dado['bacia'],
                                nom_subbacia = dado['subbacia'],
                                data_prev    = pd.to_datetime(dado['data'],format='%Y-%m-%d'),
                                num_posto    = dado['posto'],
                                val_vaz_natr = dado['vazao'],
                                val_vaz_incr = dado['vazao_inc'],
                                tempo_viagem = dado['tv'],
                                modelo       = dado['modelo'],
                                tipo_dado    = dado['tipo_dado']
                                )
                            )
            
            df_smap.apply(lambda row: escreve_arquivo(row),axis=1)
            
            #for i in range(0,len(df_smap)):
            #    
            #    vSmap.append(dict(
            #                    dat_medicao=data_arq,
            #                    nom_bacia=df_smap.loc[i,'bacia'],
            #                    nom_subbacia=df_smap.loc[i,'subbacia'],
            #                    data_prev=pd.to_datetime(df_smap.loc[i,'data'],format='%Y-%m-%d'),
            #                    num_posto=df_smap.loc[i,'posto'],
            #                    val_vaz_natr=df_smap.loc[i,'vazao'],
            #                    val_vaz_incr=df_smap.loc[i,'vazao_inc'],
            #                    tempo_viagem=df_smap.loc[i,'tv'],
            #                    modelo=df_smap.loc[i,'modelo'],
            #                    tipo_dado=df_smap.loc[i,'tipo_dado']
            #                    )
            #                )
        
        self.df_smap = pd.DataFrame(vSmap)
        
    def get_pldinter(self,plds):
        '''
        Le dados de PLD previsto pelo estudo Intersemanal
        
        :plds: Array, lista com caminhos para os arquivos que serao lidos
        '''
        
        vPld = []
        
        for pld in plds:
            
            df_pld = pd.read_csv(pld,sep=";",decimal=",")
            
            dat_medicao = datetime.strptime(pld[-12:-4],"%Y%m%d").date()
            
            for i in range(0,len(df_pld)):
                
                # Determina valor da revisao ------------------------------------------
                if df_pld.loc[i,'MEN=0-SEM=1'] == 1 and df_pld.loc[i,'Deck'][-2:] == 's1':
                
                    rev = int(df_pld.loc[i,'Deck'][-4:-3]) - 1
    
                elif df_pld.loc[i,'MEN=0-SEM=1'] == 1 and df_pld.loc[i,'Deck'][-2:] != 's1':
    
                    rev = rev + 1
                
                else:
                
                    rev = 6
                #---------------------------------------------------------------------
                
                vPld.append(dict(
                                dat_medicao   = dat_medicao,
                                modelo        = df_pld.loc[i,'Sensibilidade'],
                                nom_deck      = df_pld.loc[i,'Deck'],
                                discretizacao = df_pld.loc[i,'MEN=0-SEM=1'],
                                revisao       = int(rev),
                                pld_se        = df_pld.loc[i,'SUDESTE'],
                                pld_s         = df_pld.loc[i,'SUL'],
                                pld_ne        = df_pld.loc[i,'NORDESTE'],
                                pld_n         = df_pld.loc[i,'NORTE'],
                                )
                            )
        
        self.df_pldinter = pd.DataFrame(vPld)
        
    def get_enainter(self,enas):
        
        vEna = list()
        for ena in enas:
            
            df_enainter = pd.read_csv(ena,sep=";",decimal=",")
            df_enainter['dat_medicao'] = pd.to_datetime(df_enainter['dat_medicao'],format="%Y-%m-%d")
            
            for i,dado in df_enainter.iterrows():
                
                vEna.append(dict(
                                dat_medicao=dado.dat_medicao,
                                num_ssis=dado.num_ssis,
                                semana_op=dado.semana_op,
                                dias=dado.dias,
                                rev=dado.rev,
                                val_ena=dado.val_ena,
                                modelo=dado.modelo,
                                )
                            )
        
        
        self.df_enainter = pd.DataFrame(vEna)
        
        return
        
    def get_chuvaobs(self,obs):
        ''' Dados de chuva observada do Merge original '''
        
        vObs = list()
        for ob in obs:
            
            df_ch_obs = pd.read_csv(ob,sep=";",decimal=",")
            df_ch_obs['dat_medicao'] = pd.to_datetime(df_ch_obs['dat_medicao'],format="%Y-%m-%d")
            
            for i,dado in df_ch_obs.iterrows():
                
                vObs.append(dict(
                                dat_medicao=dado.dat_medicao,
                                nom_bacia=dado.bacia,
                                hora=dado.hora,
                                val_prec=dado.prec,
                                )
                            )
        
        
        self.df_ch_obs = pd.DataFrame(vObs)
        
        return
        
    def get_chuvaimerg(self,obs):
        ''' 
        Dados de chuva observada do IMERG original 
        
        :obs: Dataframe, lista de camnhos para os arquivos que serao lidos.
        '''
        
        vObs = list() # Inicia list que sera um Dataframe
        for ob in obs:
            
            df_ch_obs = pd.read_csv(ob,sep=";",decimal=",")                                       # Le arquivo CSV
            df_ch_obs['dat_medicao'] = pd.to_datetime(df_ch_obs['dat_medicao'],format="%Y-%m-%d") # Formata coluna de datas
            
            for i,dado in df_ch_obs.iterrows(): # Para cada linha do dataframe
                
                vObs.append(dict(
                                dat_medicao = dado.dat_medicao,
                                nom_bacia   = dado.bacia,
                                hora        = dado.hora,
                                minu        = dado.minu,
                                val_prec    = dado.prec,
                                )
                            )
        
        
        self.df_ch_obs_imerg = pd.DataFrame(vObs) # Transforma lista de dicionarios em Dataframe
        
        return
    
    def get_chuvagefs(self, gefs):
        ''' Dados do GEFS das rodadas intermediarias - 00, 06, 12 e 18z'''
        
        vGefs = list()
        
        for gf in gefs:
            
            df_ch_gefs = pd.read_csv(gf,sep=";",decimal=",")
            df_ch_gefs['dat_medicao'] = pd.to_datetime(df_ch_gefs['dat_medicao'],format="%Y-%m-%d")
            df_ch_gefs['dat_prev'] = pd.to_datetime(df_ch_gefs['dat_prev'],format="%Y-%m-%d")
            
            for i,dado in df_ch_gefs.iterrows():
                
                vGefs.append(dict(
                                dat_medicao = dado.dat_medicao,
                                dat_prev    = dado.dat_prev,
                                rodada      = dado.rodada,
                                nom_bacia   = dado.bacia,
                                val_prec    = dado.prec
                                )
                            )
        
        self.df_ch_gefs = pd.DataFrame(vGefs)
    
        return
    
    def get_carga_sem(self,cargasem):
        
        vCarga = list()
        
        for cg in cargasem:
            
            zf_file = ZipFile(cg,'r')
            zf_file.extractall(config.paths['tmp'])
            zf_filelist = zf_file.namelist()
            zf_file.close()
            
            carga_file = [os.path.join(config.paths['tmp'],x) for x in os.listdir(config.paths['tmp']) if x.find("CargaDecomp") != -1]
            
            file  = open(carga_file[0],"r")
            dados = file.readlines()
            file.close()
            
            for x in range(0,5): del dados[0]
            
            periodo = carga_file[0][(carga_file[0].find("_PMO_") + 5) : -4]
            zf_ano  = cg[-20:-18]
            meses = dict(Janeiro=1,Fevereiro=2,Março=3,Abril=4,Maio=5,Junho=6,Julho=7,Agosto=8,Setembro=9,Outubro=10,Novembro=11,Dezembro=12)
            print(carga_file[0],periodo)
            if periodo.find("Rev") == -1: # Determinando data da medicao e revisao para o caso da revisao inicial
                
                #ano      = int("20"+periodo[-2:])
                ano      = int("20"+zf_ano)
                mes_nome = periodo[:-2]
                mes      = meses[mes_nome]
                rev      = 0
                data_cg  = date(ano,mes,1)
            else: # Determinando data da medicao e revisao para o caso de revisoes posteriores a inicial
                
                #ano      = int("20"+periodo[-9:-7])
                ano      = int("20"+zf_ano)
                mes_nome = periodo[:-9]
                mes      = meses[mes_nome]
                rev      = int(periodo[-2:-1])
                data_cg  = date(ano,mes,1)
                
            for dad in dados:
                    
                aux = dad.strip().split()
                aux = [int(float(x)) for x in aux[1:]]
                
                if len(aux) > 1 and aux[1] != 11:
                
                    vCarga.append(dict(
                                dat_medicao = data_cg,# Periodo da carga
                                revisao     = rev,    # Num Revisao da carga
                                num_semana  = aux[0], # Num do estagio
                                num_ssis    = aux[1], # Num do subsistema
                                carga_p1    = aux[3], # Carga Patamar 1
                                carga_p2    = aux[5], # Carga Patamar 2
                                carga_p3    = aux[7], # Carga Patamar 3
                                hora_p1     = aux[4], # Hora Patamar 1
                                hora_p2     = aux[6], # Hora Patamar 2
                                hora_p3     = aux[8], # Hora Patamar 3
                                carga_media = round(((aux[3]*aux[4])+ \
                                (aux[5]*aux[6])+(aux[7]*aux[8]))/(aux[4]+aux[6]+aux[8]),2), # Carga Média
                            )
                        )
                
            for zf_del in zf_filelist: os.remove(os.path.join(config.paths['tmp'],zf_del)) # Apagando arquivos extraidos
            
        df_carga_sem = pd.DataFrame(vCarga)
        df_carga_sem = df_carga_sem.sort_values(['dat_medicao','revisao']).reset_index(drop=True)
                        
        self.df_carga_sem = df_carga_sem
    
    def get_vazao_sem(self, vazaosem):
        
        vDados = list()
        for file in vazaosem:
            
            zf = ZipFile(file,"r")
            
            for zfile in zf.namelist():
                
                if zfile.find("-preliminar.xls") != -1:
                    
                    zf.extract(zfile,config.paths['tmp'])
                    xlsfile = os.path.join(config.paths['tmp'],zfile)
                    zf.close()
                    break
            
            try:
                wb = pyxl.load_workbook(
                    filename=xlsfile,
                    read_only=False,
                )
            
            except:
                
                pyexcel.save_book_as(
                    file_name=xlsfile,
                    dest_file_name='{}.xlsx'.format(xlsfile[:-4])
                )
            
                wb = pyxl.load_workbook(
                    filename='{}.xlsx'.format(xlsfile[:-4]),
                    read_only=False,
                )
            

            
            if xlsfile.split("/")[-1].find("PMO") != -1:
                
                ws = wb.worksheets[3] # Aba Tab-5-6-7
                
                data_ini   = ws.cell(4,2).value
                data_fim   = ws.cell(6,2).value
                
                ena_sem_se = ws.cell(8,2).value
                ena_sem_s  = ws.cell(9,2).value
                ena_sem_ne = ws.cell(10,2).value
                ena_sem_n  = ws.cell(11,2).value
                
                ena_sem_se_perc = round(ws.cell(8,3).value/100,2)
                ena_sem_s_perc  = round(ws.cell(9,3).value/100,2)
                ena_sem_ne_perc = round(ws.cell(10,3).value/100,2)
                ena_sem_n_perc  = round(ws.cell(11,3).value/100,2)
                
                ena_mes_se = ws.cell(8,4).value
                ena_mes_s  = ws.cell(9,4).value
                ena_mes_ne = ws.cell(10,4).value
                ena_mes_n  = ws.cell(11,4).value
                
                ena_mes_se_perc = round(ws.cell(8,5).value/100,2)
                ena_mes_s_perc  = round(ws.cell(9,5).value/100,2)
                ena_mes_ne_perc = round(ws.cell(10,5).value/100,2)
                ena_mes_n_perc  = round(ws.cell(11,5).value/100,2)
                
                revisao = 0
                
                vDados.append(dict(
                                   dat_medicao     = date(data_fim.year,data_fim.month,1),
                                   data_ini        = data_ini,
                                   data_fim        = data_fim,
                                   revisao         = revisao,
                                   ena_sem_se      = ena_sem_se,
                                   ena_sem_s       = ena_sem_s,
                                   ena_sem_ne      = ena_sem_ne,
                                   ena_sem_n       = ena_sem_n,
                                   ena_sem_se_perc = ena_sem_se_perc,
                                   ena_sem_s_perc  = ena_sem_s_perc,
                                   ena_sem_ne_perc = ena_sem_ne_perc,
                                   ena_sem_n_perc  = ena_sem_n_perc,
                                   ena_mes_se      = ena_mes_se,
                                   ena_mes_s       = ena_mes_s,
                                   ena_mes_ne      = ena_mes_ne,
                                   ena_mes_n       = ena_mes_n,
                                   ena_mes_se_perc = ena_mes_se_perc,
                                   ena_mes_s_perc  = ena_mes_s_perc,
                                   ena_mes_ne_perc = ena_mes_ne_perc,
                                   ena_mes_n_perc  = ena_mes_n_perc,
                                   )
                                )
                                
            else:
                
                ws = wb.worksheets[1] # Aba REV-2
                
                data_ini   = ws.cell(5,3).value
                data_fim   = ws.cell(7,3).value
                
                ena_sem_se = ws.cell(9,3).value
                ena_sem_s  = ws.cell(10,3).value
                ena_sem_ne = ws.cell(11,3).value
                ena_sem_n  = ws.cell(12,3).value
                
                ena_sem_se_perc = round(ws.cell(9,4).value/100,2)
                ena_sem_s_perc  = round(ws.cell(10,4).value/100,2)
                ena_sem_ne_perc = round(ws.cell(11,4).value/100,2)
                ena_sem_n_perc  = round(ws.cell(12,4).value/100,2)
                
                ena_mes_se = ws.cell(9,5).value
                ena_mes_s  = ws.cell(10,5).value
                ena_mes_ne = ws.cell(11,5).value
                ena_mes_n  = ws.cell(12,5).value
                
                ena_mes_se_perc = round(ws.cell(9,6).value/100,2)
                ena_mes_s_perc  = round(ws.cell(10,6).value/100,2)
                ena_mes_ne_perc = round(ws.cell(11,6).value/100,2)
                ena_mes_n_perc  = round(ws.cell(12,6).value/100,2)
                
                revisao = int(xlsfile.split("/")[-1][35:36])
                
                vDados.append(dict(
                                   dat_medicao     = date(data_fim.year,data_fim.month,1),
                                   data_ini        = data_ini,
                                   data_fim        = data_fim,
                                   revisao         = revisao,
                                   ena_sem_se      = ena_sem_se,
                                   ena_sem_s       = ena_sem_s,
                                   ena_sem_ne      = ena_sem_ne,
                                   ena_sem_n       = ena_sem_n,
                                   ena_sem_se_perc = ena_sem_se_perc,
                                   ena_sem_s_perc  = ena_sem_s_perc,
                                   ena_sem_ne_perc = ena_sem_ne_perc,
                                   ena_sem_n_perc  = ena_sem_n_perc,
                                   ena_mes_se      = ena_mes_se,
                                   ena_mes_s       = ena_mes_s,
                                   ena_mes_ne      = ena_mes_ne,
                                   ena_mes_n       = ena_mes_n,
                                   ena_mes_se_perc = ena_mes_se_perc,
                                   ena_mes_s_perc  = ena_mes_s_perc,
                                   ena_mes_ne_perc = ena_mes_ne_perc,
                                   ena_mes_n_perc  = ena_mes_n_perc,
                                   )
                                )
            wb.close()
            shutil.rmtree(os.path.join(config.paths['tmp'],"Nao_Consistido"))
            
        df_vaz_sem = pd.DataFrame(vDados)
        self.df_vaz_sem = df_vaz_sem
        
        return
        
    def get_carga_mes(self,cargames):
        '''
        Coleta dados de carga mensal do arquivo do PMO
        
        :cargames: String, caminhos para os arquivos que serao lidos
        '''
        
        subsistemas = dict(SUDESTE=1,SUL=2,NORDESTE=3,NORTE=4) # Dicionario para conversao de subsistema em numero
        
        vCarga = list()
        for cg in cargames: # Para cada aquivo de carga
            
            # Descompacta arquivo Zip de carga ----------
            zf_file = ZipFile(cg,'r')
            zf_file.extractall(config.paths['tmp'])
            zf_filelist = zf_file.namelist()
            zf_file.close()
            #--------------------------------------------
            
            # Seleciona arquivo de carga mensal ---------------
            carga_file = [os.path.join(config.paths['tmp'],x) for x in os.listdir(config.paths['tmp']) if x.find("CargaMensal_PMO") != -1]
            #--------------------------------------------------
            
            excel_file = pd.ExcelFile(carga_file[0]) # Le arquivo Excel com Pandas
            sheet      = pd.read_excel(excel_file)   # Le dados da aba do arquivo Excel que contem a informacao de carga
            
            sheet = sheet[sheet['TYPE']=="MEDIUM"].reset_index(drop=True) # Seleciona apenas os dados de carga media
            
            for i,n in sheet.iterrows(): # Para cada linha do arquivo sheet
                
                vCarga.append(dict(
                                    data_gera = n['REVISION'],
                                    num_ssis  = subsistemas[n['SOURCE']],
                                    mes       = n['DATE'].month,
                                    ano       = n['DATE'].year,
                                    carga_med = n['LOAD']
                                    )
                                )
            
            os.remove(carga_file[0]) # Exclui o arquivo apos processamento
            
        df_carga_mes = pd.DataFrame(vCarga)                                                       # Transforma em DataFrame
        df_carga_mes = df_carga_mes.sort_values(['data_gera','mes','ano']).reset_index(drop=True) # Sort no DataFrame
                        
        self.df_carga_mes = df_carga_mes
        
        return
    
class Calculador(object):

    def realiza_calculo(self, dados, posto):
        self.vaz_calculada = posto.calcula(dados)

class Banco(object):
    def __init__(self, config):
        self.config = config
        #self.arquivos = arquivos

    def insert_rdh(self, hidro):

        session = Session(bind=engine)

        # Remocao dos dados
        min_date = pd.to_datetime(hidro.dados['dat_medicao'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.dados['dat_medicao'].unique().max()).to_pydatetime()

        session.query(Vazao). \
            filter(Vazao.dat_medicao >= min_date). \
            filter(Vazao.dat_medicao <= max_date). \
            filter(Vazao.cod_fonte == hidro.origem).delete(synchronize_session=False)

        try:
            session.commit()
            print('Dados a partir de {:%Y-%m-%d} foram deletados'.format(min_date))

        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao dos postos que ja vem no RDH
        dados = list()
        for i, dado in hidro.dados.iterrows():

            dados.append(
                Vazao(
                    num_posto=dado.num_posto,
                    dat_medicao=pd.to_datetime(dado.dat_medicao,format='%d/%m/%Y').to_pydatetime(),
                    #dat_medicao=dado.dat_medicao,
                    val_vaz_natr=dado.val_vaz_natr,
                    val_vaz_incr=dado.val_vaz_incr,
                    val_vaz_defl=dado.val_vaz_defl,
                    val_vaz_aflu=dado.val_vaz_aflu,
                    val_vaz_vert=dado.val_vaz_vert,
                    val_cota=dado.val_cota,
                    val_arm=dado.val_arm,
                    cod_fonte=hidro.origem
                )

            )

        session.bulk_save_objects(objects=dados)
        #session.add_all(dados)
        #session.commit()

        # Determina todos os postos calculados
        for posto in PostoCalc.__subclasses__():
            calculador = Calculador()
            print(posto.__name__)
            calculador.realiza_calculo(dados=hidro.dados, posto=posto())

            # verifica se posto ja existe na base
            min_date = pd.to_datetime(calculador.vaz_calculada['dat_medicao'].min()).to_pydatetime()
            max_date = pd.to_datetime(calculador.vaz_calculada['dat_medicao'].max()).to_pydatetime()
            posto = calculador.vaz_calculada['num_posto'].unique().item()

            # Deleta informacoes do posto dentro das datas do conjunto de dados passado
            session.query(Vazao). \
                filter(Vazao.dat_medicao >= min_date). \
                filter(Vazao.dat_medicao <= max_date). \
                filter(Vazao.num_posto == posto). \
                filter(Vazao.cod_fonte == hidro.origem).delete(synchronize_session=False)

            # Insercao dos dados na base
            dados = list()
            for i, dado in calculador.vaz_calculada.iterrows():

                dados.append(
                    Vazao(
                        num_posto=dado.num_posto,
                        dat_medicao=pd.to_datetime(dado.dat_medicao).to_pydatetime(),
                        val_vaz_natr=dado.val_vaz_natr,
                        cod_fonte=hidro.origem
                    )
                )
            # insercao do posto na base
            try:
                session.bulk_save_objects(objects=dados)
                #session.add_all(dados)
                session.commit()
                print('Posto calculado inserido')

            except:
                print('Posto calculado ja contido no RDH')


        session.close()
        pass

    def insert_acomph(self, hidro):

        session = Session(bind=engine)

        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.dados['dat_medicao'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.dados['dat_medicao'].unique().max()).to_pydatetime()
        
        session.query(Vazao). \
            filter(Vazao.dat_medicao >= min_date). \
            filter(Vazao.dat_medicao <= max_date). \
            filter(Vazao.cod_fonte == hidro.origem).delete(synchronize_session=False)
        try:
            session.commit()
            print('Dados a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores dos postos do acomph
        dados = list()
        for i, dado in hidro.dados.iterrows():
            # Cria objetos para insercao
            dados.append(
                Vazao(
                    num_posto=dado.num_posto,
                    dat_medicao=dado.dat_medicao.to_pydatetime(),
                    val_vaz_natr=dado.val_vaz_natr,
                    val_vaz_incr=dado.val_vaz_incr,
                    val_vaz_defl=dado.val_vaz_defl,
                    val_vaz_aflu=dado.val_vaz_aflu,
                    val_cota=dado.val_cota,
                    cod_fonte=hidro.origem
                )
            )

        #session.bulk_save_objects(objects=dados)
        #try:
        #    session.commit()
        #    
        #except:
        #    print('Nao foi.')
        
        # Determina todos os postos calculados
        for posto in PostoCalc.__subclasses__():
            calculador = Calculador()
            print(posto.__name__)
            calculador.realiza_calculo(dados=hidro.dados, posto=posto())
            
            # verifica se posto ja existe na base
            min_date = pd.to_datetime(calculador.vaz_calculada['dat_medicao'].min()).to_pydatetime()
            max_date = pd.to_datetime(calculador.vaz_calculada['dat_medicao'].max()).to_pydatetime()
            posto = calculador.vaz_calculada['num_posto'].unique().item()
            
            # Deleta informacoes do posto dentro das datas do conjunto de dados passado
            session.query(Vazao). \
                filter(Vazao.dat_medicao >= min_date). \
                filter(Vazao.dat_medicao <= max_date). \
                filter(Vazao.num_posto == posto). \
                filter(Vazao.cod_fonte == hidro.origem).delete(synchronize_session=False)
            
            #dados = list()
            for i, dado in calculador.vaz_calculada.iterrows():
                dados.append(
                    Vazao(
                        num_posto=dado.num_posto,
                        dat_medicao=dado.dat_medicao.to_pydatetime(),
                        val_vaz_natr=dado.val_vaz_natr,
                        cod_fonte=hidro.origem
                    )
                )

        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            
        except:
            print('Nao foi.')
        
        #session.close()
        pass
    
    def insert_ipdo(self, hidro):
        
        session = Session(bind=engine)
        
        # Dados gerais do IPDO ---------------------------------------------------------------
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_ipdo['dat_medicao'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.df_ipdo['dat_medicao'].unique().max()).to_pydatetime()
        
        session.query(Ipdo). \
            filter(Ipdo.dat_medicao >= min_date.date()). \
            filter(Ipdo.dat_medicao <= max_date.date()).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados do IPDO a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))
        
        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_ipdo.iterrows():
            # Cria objetos para insercao
            dados.append(
                Ipdo(
                    dat_medicao=pd.to_datetime(dado.dat_medicao,format='%Y-%m-%d').to_pydatetime(),
                    num_ssis=dado.num_ssis,
                    nom_ssis=dado.nom_ssis,
                    ena_ssis=dado.ena_ssis,
                    arm_ssis=dado.arm_ssis,
                    carga_ssis=dado.carga_ssis
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do IPDO inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        # Armazenamento individual ----------------------------------------------------------
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_arm_ipdo['dat_medicao'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.df_arm_ipdo['dat_medicao'].unique().max()).to_pydatetime()
        
        session.query(ArmIpdo). \
            filter(ArmIpdo.dat_medicao >= min_date.date()). \
            filter(ArmIpdo.dat_medicao <= max_date.date()).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados do de arm. individualizado do IPDO a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))
        
        dados = list()
        for i, dado in hidro.df_arm_ipdo.iterrows():
            # Cria objetos para insercao
            dados.append(
                ArmIpdo(
                    dat_medicao=pd.to_datetime(dado.dat_medicao,format='%Y-%m-%d').to_pydatetime(),
                    origem=dado.origem,
                    val_arm=dado.val_arm,
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de arm. individualizado do IPDO inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
        
    def insert_pmedia(self, hidro,modelo):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_ch['data_gera'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.df_ch['data_gera'].unique().max()).to_pydatetime()
       
        session.query(Previsao). \
            filter(Previsao.data_gera >= min_date). \
            filter(Previsao.data_gera <= max_date). \
            filter(Previsao.modelo == modelo).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados de previsao do modelo '+modelo+' a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_ch.iterrows():
            # Cria objetos para insercao
            #print("Lendo dados de "+str(dado.data_gera)+".")
            dados.append(
                Previsao(
                    data_gera=dado.data_gera,
                    data_prev=dado.data_prev,
                    lon=dado.lon,
                    lat=dado.lat,
                    prec=dado.prec,
                    modelo=dado.modelo
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de previsao do modelo "+modelo+" inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_merge(self, hidro):
        
        if hidro.df_merge.empty:
            return -1
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_merge['data_obs'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.df_merge['data_obs'].unique().max()).to_pydatetime()
       
        session.query(Merge). \
            filter(Merge.data_obs >= min_date). \
            filter(Merge.data_obs <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados do merge a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_merge.iterrows():
            # Cria objetos para insercao
            #print("Lendo dados de "+str(dado.data_gera)+".")
            dados.append(
                Merge(
                    nom_bacia=dado.nom_bacia,
                    data_obs=dado.data_obs,
                    prec=dado.prec
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do merge inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_smap_ec45(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().max())#.to_pydatetime()

        session.query(ChuvaVazaoEc45). \
            filter(ChuvaVazaoEc45.dat_medicao >= min_date). \
            filter(ChuvaVazaoEc45.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados do chuva-vazao com EC45 a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_smap.iterrows():
            # Cria objetos para insercao
            dados.append(
                ChuvaVazaoEc45(
                    dat_medicao=dado.dat_medicao.to_pydatetime(),
                    nom_bacia=dado.nom_bacia,
                    nom_subbacia=dado.nom_subbacia,
                    data_prev=dado.data_prev.to_pydatetime(),
                    num_posto=dado.num_posto,
                    val_vaz_natr=dado.val_vaz_natr,
                    val_vaz_incr=dado.val_vaz_incr,
                    tempo_viagem=dado.tempo_viagem,
                    modelo=dado.modelo,
                    tipo_dado=dado.tipo_dado
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do chuva-vazao inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
        
    def insert_smap(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().max())#.to_pydatetime()

        session.query(ChuvaVazao). \
            filter(ChuvaVazao.dat_medicao >= min_date). \
            filter(ChuvaVazao.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados do chuva-vazao a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_smap.iterrows():
            # Cria objetos para insercao
            dados.append(
                ChuvaVazao(
                    dat_medicao=dado.dat_medicao.to_pydatetime(),
                    nom_bacia=dado.nom_bacia,
                    nom_subbacia=dado.nom_subbacia,
                    data_prev=dado.data_prev.to_pydatetime(),
                    num_posto=dado.num_posto,
                    val_vaz_natr=dado.val_vaz_natr,
                    val_vaz_incr=dado.val_vaz_incr,
                    tempo_viagem=dado.tempo_viagem,
                    modelo=dado.modelo,
                    tipo_dado=dado.tipo_dado
                )
            )
        #for i in range(0,len(hidro.df_smap)):
        #    # Cria objetos para insercao
        #    dados.append(
        #        ChuvaVazao(
        #            dat_medicao  = hidro.df_smap.loc[i,'dat_medicao'].to_pydatetime(),
        #            nom_bacia    = hidro.df_smap.loc[i,'nom_bacia'],
        #            nom_subbacia = hidro.df_smap.loc[i,'nom_subbacia'],
        #            data_prev    = hidro.df_smap.loc[i,'data_prev'].to_pydatetime(),
        #            num_posto    = hidro.df_smap.loc[i,'num_posto'],
        #            val_vaz_natr = hidro.df_smap.loc[i,'val_vaz_natr'],
        #            val_vaz_incr = hidro.df_smap.loc[i,'val_vaz_incr'],
        #            tempo_viagem = hidro.df_smap.loc[i,'tempo_viagem'],
        #            modelo       = hidro.df_smap.loc[i,'modelo'],
        #            tipo_dado    = hidro.df_smap.loc[i,'tipo_dado'],
        #        )
        #    )
        
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do chuva-vazao inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_smap_hist(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_smap['dat_medicao'].unique().max())#.to_pydatetime()

        session.query(ChuvaVazaoHist). \
            filter(ChuvaVazaoHist.dat_medicao >= datetime(2019,1,1)). \
            filter(ChuvaVazaoHist.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados do chuva-vazao (Historico) a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))
        
        dados = list()
        def escreve_arquivo(dado):
            
            dados.append(
                ChuvaVazaoHist(
                    dat_medicao  = dado.dat_medicao.to_pydatetime(),
                    nom_bacia    = dado.nom_bacia,
                    nom_subbacia = dado.nom_subbacia,
                    data_prev    = dado.data_prev.to_pydatetime(),
                    num_posto    = dado.num_posto,
                    val_vaz_natr = dado.val_vaz_natr,
                    val_vaz_incr = dado.val_vaz_incr,
                    tempo_viagem = dado.tempo_viagem,
                    modelo       = dado.modelo,
                    tipo_dado    = dado.tipo_dado
                )
            )
        
        hidro.df_smap.apply(lambda row: escreve_arquivo(row),axis=1)
        
        ## Insercao / atualizacao dos valores
        #dados = list()
        #for i, dado in hidro.df_smap.iterrows():
        #    # Cria objetos para insercao
        #    dados.append(
        #        ChuvaVazaoHist(
        #            dat_medicao  = dado.dat_medicao.to_pydatetime(),
        #            nom_bacia    = dado.nom_bacia,
        #            nom_subbacia = dado.nom_subbacia,
        #            data_prev    = dado.data_prev.to_pydatetime(),
        #            num_posto    = dado.num_posto,
        #            val_vaz_natr = dado.val_vaz_natr,
        #            val_vaz_incr = dado.val_vaz_incr,
        #            tempo_viagem = dado.tempo_viagem,
        #            modelo       = dado.modelo,
        #            tipo_dado    = dado.tipo_dado
        #        )
        #    )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do chuva-vazao (Historico) inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_pldinter(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data 
        min_date = pd.to_datetime(hidro.df_pldinter['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_pldinter['dat_medicao'].unique().max())#.to_pydatetime()
       
        session.query(PldInter). \
            filter(PldInter.dat_medicao >= min_date). \
            filter(PldInter.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados de PLD do intersemanal a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i in range(0,len(hidro.df_pldinter)):
            # Cria objetos para insercao
            dados.append(
                        PldInter(
                                dat_medicao   = hidro.df_pldinter.loc[i,'dat_medicao'],
                                modelo        = hidro.df_pldinter.loc[i,'modelo'],
                                nom_deck      = hidro.df_pldinter.loc[i,'nom_deck'],
                                discretizacao = int(hidro.df_pldinter.loc[i,'discretizacao']),
                                revisao       = int(hidro.df_pldinter.loc[i,'revisao']),
                                pld_se        = hidro.df_pldinter.loc[i,'pld_se'],
                                pld_s         = hidro.df_pldinter.loc[i,'pld_s'],
                                pld_ne        = hidro.df_pldinter.loc[i,'pld_ne'],
                                pld_n         = hidro.df_pldinter.loc[i,'pld_n'],
                                )
                        )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de PLD do intersemanal inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_enainter(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data 
        min_date = pd.to_datetime(hidro.df_enainter['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_enainter['dat_medicao'].unique().max())#.to_pydatetime()
       
        session.query(EnaInter). \
            filter(EnaInter.dat_medicao >= min_date). \
            filter(EnaInter.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados de ENA do intersemanal a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_enainter.iterrows():
            # Cria objetos para insercao
            dados.append(
                EnaInter(
                        dat_medicao=dado.dat_medicao,
                        num_ssis=dado.num_ssis,
                        semana_op=dado.semana_op,
                        dias=dado.dias,
                        rev=dado.rev,
                        val_ena=dado.val_ena,
                        modelo=dado.modelo,
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de ENA do intersemanal inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_enainter_ec45(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data 
        min_date = pd.to_datetime(hidro.df_enainter['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_enainter['dat_medicao'].unique().max())#.to_pydatetime()
       
        session.query(EnaInterEc45). \
            filter(EnaInterEc45.dat_medicao >= min_date). \
            filter(EnaInterEc45.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados de ENA do intersemanal (com ECMWF de 45 dias) a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_enainter.iterrows():
            # Cria objetos para insercao
            dados.append(
                EnaInterEc45(
                        dat_medicao=dado.dat_medicao,
                        num_ssis=dado.num_ssis,
                        semana_op=dado.semana_op,
                        dias=dado.dias,
                        rev=dado.rev,
                        val_ena=dado.val_ena,
                        modelo=dado.modelo,
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de ENA do intersemanal inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_chuvaobs(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data
        min_date = pd.to_datetime(hidro.df_ch_obs['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_ch_obs['dat_medicao'].unique().max())#.to_pydatetime()
        
        min_hora_min = int(hidro.df_ch_obs[hidro.df_ch_obs['dat_medicao']==min_date]['hora'].unique().min())
        max_hora_min = int(hidro.df_ch_obs[hidro.df_ch_obs['dat_medicao']==min_date]['hora'].unique().max())
        
        min_hora_max = int(hidro.df_ch_obs[hidro.df_ch_obs['dat_medicao']==max_date]['hora'].unique().min())
        max_hora_max = int(hidro.df_ch_obs[hidro.df_ch_obs['dat_medicao']==max_date]['hora'].unique().max())

        session.query(ChuvaObs). \
            filter(ChuvaObs.dat_medicao == min_date). \
            filter(ChuvaObs.hora >= min_hora_min). \
            filter(ChuvaObs.hora <= max_hora_min).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados de chuva observada a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
            
        session.query(ChuvaObs). \
            filter(ChuvaObs.dat_medicao == max_date). \
            filter(ChuvaObs.hora >= min_hora_max). \
            filter(ChuvaObs.hora <= max_hora_max).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados de chuva observada a partir de {:%Y-%m-%d} foram deletados'.format(max_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(max_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_ch_obs.iterrows():
            # Cria objetos para insercao
            dados.append(
                ChuvaObs(
                        dat_medicao = dado.dat_medicao,
                        nom_bacia   = dado.nom_bacia,
                        hora        = dado.hora,
                        val_prec    = dado.val_prec
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de chuva observada inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        return
        
    def insert_chuvaimerg(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data
        # Datas maximas e minimas ----------------------------------
        min_date = pd.to_datetime(hidro.df_ch_obs_imerg['dat_medicao'].unique().min()) # Data minima dos dados
        max_date = pd.to_datetime(hidro.df_ch_obs_imerg['dat_medicao'].unique().max()) # Data maxima dos dados
        #-----------------------------------------------------------
        
        # Horas minimas e maximas ----------------------------------
        min_hora_min = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==min_date]['hora'].unique().min()) # Hora minima do menor dia
        max_hora_min = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==min_date]['hora'].unique().max()) # Hora maxima do menor dia
        
        min_hora_max = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==max_date]['hora'].unique().min()) # Hora minima do maior dia
        max_hora_max = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==max_date]['hora'].unique().max()) # Hora maxima do maior dia
        #-----------------------------------------------------------
        
        # Minutos minimos e maximos --------------------------------
        min_minu_min = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==min_date]['minu'].unique().min()) # Hora minima do menor dia
        max_minu_min = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==min_date]['minu'].unique().max()) # Hora maxima do menor dia
        
        min_minu_max = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==max_date]['minu'].unique().min()) # Hora minima do maior dia
        max_minu_max = int(hidro.df_ch_obs_imerg[hidro.df_ch_obs_imerg['dat_medicao']==max_date]['minu'].unique().max()) # Hora maxima do maior dia
        #-----------------------------------------------------------
        
        # Remove todos os dados a partir da primeira data, hora e minuto --------
        session.query(ChuvaObsImerg). \
            filter(ChuvaObsImerg.dat_medicao == min_date). \
            filter(ChuvaObsImerg.hora >= min_hora_min). \
            filter(ChuvaObsImerg.hora <= max_hora_min). \
            filter(ChuvaObsImerg.minu >= min_minu_min). \
            filter(ChuvaObsImerg.minu <= max_minu_min).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados de chuva observada (IMERG) a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
            
        session.query(ChuvaObsImerg). \
            filter(ChuvaObsImerg.dat_medicao == max_date). \
            filter(ChuvaObsImerg.hora >= min_hora_max). \
            filter(ChuvaObsImerg.hora <= max_hora_max). \
            filter(ChuvaObsImerg.minu >= min_minu_max). \
            filter(ChuvaObsImerg.minu <= max_minu_max).delete(synchronize_session=False)
        
        try:
            session.commit()
            print('Dados de chuva observada (IMERG) a partir de {:%Y-%m-%d} foram deletados'.format(max_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(max_date))
        #------------------------------------------------------------------------
        
        
        # Insercao / atualizacao dos valores ------------------------------------
        dados = list()
        for i, dado in hidro.df_ch_obs_imerg.iterrows():
            # Cria objetos para insercao
            dados.append(
                ChuvaObsImerg(
                        dat_medicao = dado.dat_medicao,
                        nom_bacia   = dado.nom_bacia,
                        hora        = dado.hora,
                        minu        = dado.minu,
                        val_prec    = dado.val_prec
                )
            )
        #------------------------------------------------------------------------
        
        # Insercao em massa -----------------------------------------------------
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de chuva observada inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        #------------------------------------------------------------------------
        
        return
    
    def insert_gefs(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data 
        min_date = pd.to_datetime(hidro.df_ch_gefs['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_ch_gefs['dat_medicao'].unique().max())#.to_pydatetime()
        
        min_rod_min = int(hidro.df_ch_gefs[hidro.df_ch_gefs['dat_medicao']==min_date]['rodada'].unique().min())
        max_rod_min = int(hidro.df_ch_gefs[hidro.df_ch_gefs['dat_medicao']==min_date]['rodada'].unique().max())
        
        min_rod_max = int(hidro.df_ch_gefs[hidro.df_ch_gefs['dat_medicao']==max_date]['rodada'].unique().min())
        max_rod_max = int(hidro.df_ch_gefs[hidro.df_ch_gefs['dat_medicao']==max_date]['rodada'].unique().max())

        session.query(ChuvaGefs). \
            filter(ChuvaGefs.dat_medicao == min_date). \
            filter(ChuvaGefs.rodada >= min_rod_min). \
            filter(ChuvaGefs.rodada <= max_rod_min).delete(synchronize_session=False)

        try:
            session.commit()
            print('Dados de chuva do GEFS de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
        
        if min_date != max_date:
            session.query(ChuvaGefs). \
                filter(ChuvaGefs.dat_medicao == max_date). \
                filter(ChuvaGefs.rodada >= min_rod_max). \
                filter(ChuvaGefs.rodada <= max_rod_max).delete(synchronize_session=False)
            
            try:
                session.commit()
                print('Dados de chuva do GEFS a partir de {:%Y-%m-%d} foram deletados'.format(max_date))
            except:
                print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(max_date))
        
            #session.query(ChuvaGefs). \
            #    filter(ChuvaGefs.dat_medicao >= min_date). \
            #    filter(ChuvaGefs.dat_medicao <= max_date).delete(synchronize_session=False)

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_ch_gefs.iterrows():
            # Cria objetos para insercao
            dados.append(
                ChuvaGefs(
                        dat_medicao = dado.dat_medicao,
                        dat_prev    = dado.dat_prev,
                        rodada      = dado.rodada,
                        nom_bacia   = dado.nom_bacia,
                        val_prec    = dado.val_prec
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de chuva do modelo GEFS inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_carga_sem(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da referencia de data e revisao
        min_date = pd.to_datetime(hidro.df_carga_sem['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_carga_sem['dat_medicao'].unique().max())#.to_pydatetime()
        
        revisoes_min = hidro.df_carga_sem[hidro.df_carga_sem['dat_medicao']==min_date]['revisao'].unique() # Revisoes da data minima do dataframe
        revisoes_max = hidro.df_carga_sem[hidro.df_carga_sem['dat_medicao']==max_date]['revisao'].unique() # Revisoes da data maxima do dataframe

        for rev in revisoes_min:

            session.query(CargaSem). \
                filter(CargaSem.data_gera == min_date). \
                filter(CargaSem.revisao == int(rev)).delete(synchronize_session=False)
            
            try:
                session.commit()
                print('Dados de carga semanal de {:%Y-%m-%d} revisao {} foram deletados'.format(min_date,rev))
            except:
                print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
                
        if min_date != max_date: # Caso o dataframe possua dados de carga de dois meses diferentes
        
            for rev in revisoes_max:
            
                session.query(CargaSem). \
                    filter(CargaSem.data_gera == max_date). \
                    filter(CargaSem.revisao == int(rev)).delete(synchronize_session=False)
                    
                try:
                    session.commit()
                    print('Dados de carga semanal de {:%Y-%m-%d} revisao {} foram deletados'.format(max_date,rev))
                except:
                    print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(max_date))
        
        # Insercao / atualizacao dos valores
        dados = list()
        
        for i, dado in hidro.df_carga_sem.iterrows():
            # Cria objetos para insercao
            dados.append(
                CargaSem(
                        data_gera   = dado.dat_medicao, # Data de referencia
                        revisao     = dado.revisao,     # Numero da revisao
                        num_semana  = dado.num_semana,  # Numero do estagio
                        num_ssis    = dado.num_ssis,    # Numedo do subsistema
                        carga_p1    = dado.carga_p1,    # Carga Patamar 1
                        carga_p2    = dado.carga_p2,    # Carga Patamar 2
                        carga_p3    = dado.carga_p3,    # Carga Patamar 3
                        hora_p1     = dado.hora_p1,     # Hora Patamar 1
                        hora_p2     = dado.hora_p2,     # Hora Patamar 2
                        hora_p3     = dado.hora_p3,     # Hora Patamar 3
                        carga_med   = dado.carga_media, # Carga media do estagio
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de carga semanal inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
        
        
    def insert_vazao_sem(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da referencia de data e revisao
        min_date = pd.to_datetime(hidro.df_vaz_sem['dat_medicao'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_vaz_sem['dat_medicao'].unique().max())#.to_pydatetime()
        
        revisoes_min = hidro.df_vaz_sem[hidro.df_vaz_sem['dat_medicao']==min_date]['revisao'].unique() # Revisoes da data minima do dataframe
        revisoes_max = hidro.df_vaz_sem[hidro.df_vaz_sem['dat_medicao']==max_date]['revisao'].unique() # Revisoes da data maxima do dataframe

        for rev in revisoes_min:

            session.query(VazaoSem). \
                filter(VazaoSem.dat_medicao == min_date). \
                filter(VazaoSem.revisao == int(rev)).delete(synchronize_session=False)
            
            try:
                session.commit()
                print('Dados de vazoes semanais de {:%Y-%m-%d} revisao {} foram deletados'.format(min_date,rev))
            except:
                print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
                
        if min_date != max_date: # Caso o dataframe possua dados de carga de dois meses diferentes
        
            for rev in revisoes_max:
            
                session.query(VazaoSem). \
                    filter(VazaoSem.dat_medicao == max_date). \
                    filter(VazaoSem.revisao == int(rev)).delete(synchronize_session=False)
                    
                try:
                    session.commit()
                    print('Dados de vazoes semanais de {:%Y-%m-%d} revisao {} foram deletados'.format(max_date,rev))
                except:
                    print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(max_date))
                    
         # Insercao / atualizacao dos valores
        dados = list()
        
        for i, dado in hidro.df_vaz_sem.iterrows():
            # Cria objetos para insercao
            dados.append(
                VazaoSem(
                        dat_medicao     = dado.dat_medicao,     # Data de referencia
                        data_ini        = dado.data_ini,        # Data inicio da semana operativa
                        data_fim        = dado.data_fim,        # Data fim da semana operativa
                        revisao         = dado.revisao,         # Numero da revisao do mes
                        ena_sem_se      = dado.ena_sem_se,      # ENA media semanal do SE
                        ena_sem_s       = dado.ena_sem_s,       # ENA media semanal do S
                        ena_sem_ne      = dado.ena_sem_ne,      # ENA media semanal do NE
                        ena_sem_n       = dado.ena_sem_n,       # ENA media semanal do N
                        ena_sem_se_perc = dado.ena_sem_se_perc, # ENA media semanal percentual SE
                        ena_sem_s_perc  = dado.ena_sem_s_perc,  # ENA media semanal percentual S
                        ena_sem_ne_perc = dado.ena_sem_ne_perc, # ENA media semanal percentual NE
                        ena_sem_n_perc  = dado.ena_sem_n_perc,  # ENA media semanal percentual N
                        ena_mes_se      = dado.ena_mes_se,      # ENA media mensal do SE
                        ena_mes_s       = dado.ena_mes_s,       # ENA media mensal do S
                        ena_mes_ne      = dado.ena_mes_ne,      # ENA media mensal do NE
                        ena_mes_n       = dado.ena_mes_n,       # ENA media mensal do N
                        ena_mes_se_perc = dado.ena_mes_se_perc, # ENA media mensal percentual SE
                        ena_mes_s_perc  = dado.ena_mes_s_perc,  # ENA media mensal percentual S
                        ena_mes_ne_perc = dado.ena_mes_ne_perc, # ENA media mensal percentual NE
                        ena_mes_n_perc  = dado.ena_mes_n_perc,  # ENA media mensal percentual N
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de vazoes semanais inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def insert_carga_mes(self,hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da referencia de data e revisao
        min_date = pd.to_datetime(hidro.df_carga_mes['data_gera'].unique().min())#.to_pydatetime()
        max_date = pd.to_datetime(hidro.df_carga_mes['data_gera'].unique().max())#.to_pydatetime()

        session.query(CargaMes). \
            filter(CargaMes.data_gera >= min_date). \
            filter(CargaMes.data_gera <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados de carga mensal de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados de {:%Y-%m-%d}'.format(min_date))
                

        # Insercao / atualizacao dos valores
        dados = list()
        
        for i, dado in hidro.df_carga_mes.iterrows():
            # Cria objetos para insercao
            dados.append(
                CargaMes(
                        data_gera   = dado.data_gera,   # Data de referencia
                        num_ssis    = dado.num_ssis,    # Numedo do subsistema
                        mes         = dado.mes,         # Mes previsao
                        ano         = dado.ano,         # Ano 
                        carga_med   = dado.carga_med,   # Carga Media
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados de carga mensal inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        return
    
    def insert_previvaz(self, hidro):
        
        session = Session(bind=engine)
        
        # Remove todos os dados a partir da primeira data do acomph
        min_date = pd.to_datetime(hidro.df_hist['dat_medicao'].unique().min()).to_pydatetime()
        max_date = pd.to_datetime(hidro.df_hist['dat_medicao'].unique().max()).to_pydatetime()
       
        session.query(HistPvv). \
            filter(HistPvv.dat_medicao >= datetime(2018,1,1)). \
            filter(HistPvv.dat_medicao <= max_date).delete(synchronize_session=False)
            
        try:
            session.commit()
            print('Dados do historico de vazoes do previvaz a partir de {:%Y-%m-%d} foram deletados'.format(min_date))
        except:
            print('Nao foi possivel a remocao dos dados a partir de {:%Y-%m-%d}'.format(min_date))

        # Insercao / atualizacao dos valores
        dados = list()
        for i, dado in hidro.df_hist.iterrows():
            # Cria objetos para insercao

            dados.append(
                HistPvv(
                    dat_medicao = dado.dat_medicao,
                    num_posto   = dado.num_posto,
                    num_semana  = dado.num_semana,
                    num_ano     = dado.num_ano,
                    val_vaz     = dado.val_vaz,
                )
            )
        
        # Insercao em massa
        session.bulk_save_objects(objects=dados)
        try:
            session.commit()
            print("Dados do historico de vazoes do previvaz inseridos com sucesso.")
        except:
            print('Nao foi possivel inserir os dados na base.')
        
        pass
    
    def query_ena(self, data_inicial='2015-01-01', data_final='2015-12-31'):

        session = Session(bind=engine)

        t1 = aliased(Vazao, name='t1')

        stmt = session.query(
            t1.dat_medicao,
            Subsistema.num_ssis,
            Subsistema.nom_ssis,
            Bacia.num_bacia,
            Bacia.nom_bacia,
            t1.num_posto,
            func.max(t1.cod_fonte).label("cod_fonte"),
            Posto.bin_ena,
            Posto.val_produt,
            t1.val_vaz_natr,
            (Posto.val_produt * t1.val_vaz_natr).label('val_ena'),
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(Posto.bin_ena == 1). \
            join(Posto). \
            join(Bacia). \
            join(Subsistema). \
            group_by('num_ssis', 'num_bacia', 'num_posto', 'dat_medicao')

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        self.dados = pd.DataFrame(df)
        
        self.dados_sub = pd.DataFrame(df.groupby(by=['num_ssis', 'nom_ssis','dat_medicao']).agg(dict(val_ena='sum')))
        self.dados_bac = pd.DataFrame(df.groupby(by=['num_bacia', 'nom_bacia','dat_medicao']).agg(dict(val_ena='sum')))
        self.dados_posto = self.dados
        
        stmt = session.query(
            t1.dat_medicao,
            Subsistema.num_ssis,
            Subsistema.nom_ssis,
            Bacia.num_bacia,
            Bacia.nom_bacia,
            t1.num_posto,
            func.max(t1.cod_fonte).label("cod_fonte"),
            Posto.bin_ena,
            Posto.val_produt,
            t1.val_vaz_natr,
            t1.val_vaz_incr,
            (Posto.val_produt * t1.val_vaz_natr).label('val_ena'),
            (Posto.val_produt * t1.val_vaz_incr).label('val_ena_incr'),
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(t1.num_posto == 266). \
            join(Posto). \
            join(Bacia). \
            join(Subsistema). \
            group_by('num_ssis', 'num_bacia', 'num_posto', 'dat_medicao')

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        self.dados_inc = pd.DataFrame(df)
        
    def query_vazao_diaria(self, data_inicial='2015-01-01', data_final='2015-12-31'):
        
        session = Session(bind=engine)

        t1 = aliased(Vazao, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            Subsistema.num_ssis,
            t1.num_posto,
            func.max(t1.cod_fonte).label("cod_fonte"),
            Posto.bin_ena,
            Posto.cod_model,
            Posto.cod_model_2,
            Posto.val_produt,
            t1.val_vaz_natr,
            t1.val_vaz_incr,
            (Posto.val_produt * t1.val_vaz_natr).label('val_ena')
            ).\
            filter(t1.dat_medicao.between(data_inicial, data_final)).\
            join(Posto).\
            join(Subsistema).\
            group_by(t1.dat_medicao,t1.num_posto)
        
        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        self.dados_dias = pd.DataFrame(df).fillna(0)
        
        stmt = session.query(
            t1.dat_medicao,
            t1.num_posto,
            t1.cod_fonte,
            t1.val_cota,
            t1.val_arm
            ).\
            filter(t1.dat_medicao.between(data_inicial, data_final)).\
            filter(t1.cod_fonte==1).\
            group_by(t1.dat_medicao,t1.num_posto)
            
        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        self.dados_arm = pd.DataFrame(df).fillna(0)
        
        
    def query_chuva(self,data_inicial='2019-01-01', data_final='2019-12-31',modelo='eta+gefs'):
        
        session = Session(bind=engine)
        
        #if modelo == "eta+gefs":
        t1 = aliased(Grade, name='t1')
        #elif modelo == "gefs":
        #    t1 = aliased(GradeGefs, name='t1')
        #else:
        #    t1 = aliased(GradeEcmwf, name='t1')
            
        t2 = aliased(Previsao, name='t2')

        stmt = session.query(
            t1.num_bacia,
            t1.nom_bacia,
            t1.nom_subbac,
            t2.data_gera,
            t2.data_prev,
            t2.lon,
            t2.lat,
            t2.prec,
            t2.modelo,
        ). \
            filter(t2.data_gera >= data_inicial). \
            filter(t2.data_gera <= data_final). \
            filter(t2.lon == t1.lon). \
            filter(t2.lat == t1.lat). \
            filter(t2.modelo == modelo). \
            order_by('nom_bacia', 'data_gera', 'data_prev')

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados = pd.DataFrame(df)
        self.dados_ch = pd.DataFrame(df.groupby(by=['nom_bacia', 'data_gera','data_prev','modelo']).agg(dict(prec='mean')))
        
    def query_merge(self,data_inicial='2019-01-01', data_final='2019-12-31'):
        
        session = Session(bind=engine)

        t1 = aliased(Merge, name='t1')
        #t2 = aliased(Previsao, name='t2')

        stmt = session.query(
            t1.nom_bacia,
            t1.data_obs,
            t1.prec
        ). \
            filter(t1.data_obs >= data_inicial). \
            filter(t1.data_obs <= data_final). \
            order_by('data_obs','nom_bacia')

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_merge = pd.DataFrame(df)
        self.dados_merge = pd.DataFrame(df.groupby(by=['nom_bacia', 'data_obs']).agg(dict(prec='sum')))
        
    def query_ipdo(self,data_inicial='2019-01-01', data_final='2019-12-31'):
        
        session = Session(bind=engine)

        t1 = aliased(Ipdo, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.num_ssis,
            t1.nom_ssis,
            t1.ena_ssis,
            t1.arm_ssis,
            t1.carga_ssis
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            order_by('dat_medicao','num_ssis')

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_ipdo = pd.DataFrame(df)
        #self.dados_ipdo = pd.DataFrame(df.groupby(by=['dat_medicao','num_ssis']).agg(dict(ena_ssis='sum',arm_ssis='sum',carga_ssis='sum')))
    
    def query_arm_ipdo(self,data_inicial='2019-01-01', data_final='2019-12-31'):
        
        session = Session(bind=engine)
        
        t1 = aliased(ArmIpdo, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.origem,
            t1.val_arm
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            order_by('dat_medicao')
            
        df = pd.read_sql(sql=stmt.statement, con=session.bind)
       
        self.dados_arm_ipdo = pd.DataFrame(df)
        
    
    def query_smap(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazao, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.nom_bacia,
            t1.nom_subbacia,
            t1.data_prev,
            t1.num_posto,
            t1.val_vaz_natr,
            t1.val_vaz_incr,
            t1.tempo_viagem,
            t1.modelo,
            t1.tipo_dado,
            Posto.val_produt,
            #(Posto.val_produt * t1.val_vaz_natr).label('val_ena')
            
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            join(Posto). \
            order_by('dat_medicao','num_posto')

        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))
        
        #modelos = df['modelo'].unique()
        #max_dat = df['dat_medicao'].max()
        #min_dat = df['dat_medicao'].min()
        #
        #df_calc = pd.DataFrame([])
        #for dt in [min_dat, max_dat]:
        #    
        #    for mod in modelos:
        #        
        #        df_aux = df[(df['modelo']==mod) & (df['dat_medicao']==dt)].reset_index(drop=True)
        #        
        #        for posto in PostoSmap.__subclasses__():
        #            
        #            posto.calcula(PostoSmap,df_aux,mod)
        #            
        #            df_calc = pd.concat([df_calc,posto.vaz_calculada],ignore_index=True)
        #
        #df  = pd.concat([df,df_calc],ignore_index=True)
        df['val_ena'] = df['val_vaz_natr'] * df['val_produt']
        df.drop('val_produt',axis=1,inplace=True)
        
        self.dados_smap = df
        
    def query_smap_ena(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazao, name='t1')
        t2 = aliased(Posto,name='t2')
        
        stmt = session.query(
            t1.dat_medicao,
            t2.num_ssis,
            t2.num_bacia,
            t1.num_posto,
            t1.data_prev,
            t1.val_vaz_natr,
            t1.modelo,
            t2.val_produt,
            t2.bin_ena,
            (t2.val_produt * t1.val_vaz_natr).label('val_ena')
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(t2.bin_ena == 1). \
            join(t2). \
            order_by('dat_medicao','num_posto')
        
        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))
        
        df_ena = df.groupby(['dat_medicao','num_ssis','num_bacia','data_prev','modelo']).agg(dict(val_ena='sum')).reset_index()
        
        self.dados_ena_smap = df_ena
        
    def query_smap_ec45(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazaoEc45, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.nom_bacia,
            t1.nom_subbacia,
            t1.data_prev,
            t1.num_posto,
            t1.val_vaz_natr,
            t1.val_vaz_incr,
            t1.tempo_viagem,
            t1.modelo,
            t1.tipo_dado,
            Posto.val_produt,
            #(Posto.val_produt * t1.val_vaz_natr).label('val_ena')
            
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            join(Posto). \
            order_by('dat_medicao','num_posto')

        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))
        
        #modelos = df['modelo'].unique()
        #max_dat = df['dat_medicao'].max()
        #min_dat = df['dat_medicao'].min()
        #
        #df_calc = pd.DataFrame([])
        #for dt in [min_dat, max_dat]:
        #    
        #    for mod in modelos:
        #        
        #        df_aux = df[(df['modelo']==mod) & (df['dat_medicao']==dt)].reset_index(drop=True)
        #        
        #        for posto in PostoSmap.__subclasses__():
        #            
        #            posto.calcula(PostoSmap,df_aux,mod)
        #            
        #            df_calc = pd.concat([df_calc,posto.vaz_calculada],ignore_index=True)
        #
        #df  = pd.concat([df,df_calc],ignore_index=True)
        df['val_ena'] = df['val_vaz_natr'] * df['val_produt']
        df.drop('val_produt',axis=1,inplace=True)
        
        self.dados_smap_ec45 = df
    
    def query_smap_ec45_ena(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazaoEc45, name='t1')
        t2 = aliased(Posto,name='t2')
        
        stmt = session.query(
            t1.dat_medicao,
            t2.num_ssis,
            t2.num_bacia,
            t1.num_posto,
            t1.data_prev,
            t1.val_vaz_natr,
            t1.modelo,
            t2.val_produt,
            t2.bin_ena,
            (t2.val_produt * t1.val_vaz_natr).label('val_ena')
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(t2.bin_ena == 1). \
            join(t2). \
            order_by('dat_medicao','num_posto')
        
        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))
        
        df_ena = df.groupby(['dat_medicao','num_ssis','num_bacia','data_prev','modelo']).agg(dict(val_ena='sum')).reset_index()
        
        self.dados_ena_smap_ec45 = df_ena
    
    def query_smap_hist(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazaoHist, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.nom_bacia,
            t1.nom_subbacia,
            t1.data_prev,
            t1.num_posto,
            t1.val_vaz_natr,
            t1.val_vaz_incr,
            t1.tempo_viagem,
            t1.modelo,
            t1.tipo_dado,
            Posto.val_produt,
            
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            join(Posto). \
            order_by('dat_medicao','num_posto')

        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))

        df['val_ena'] = df['val_vaz_natr'] * df['val_produt']
        df.drop('val_produt',axis=1,inplace=True)
        
        self.dados_smap_hist = df
        
    def query_smap_hist_ena(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaVazaoHist, name='t1')
        t2 = aliased(Posto,name='t2')
        
        stmt = session.query(
            t1.dat_medicao,
            t2.num_ssis,
            t2.num_bacia,
            t1.num_posto,
            t1.data_prev,
            t1.val_vaz_natr,
            t1.modelo,
            t2.val_produt,
            t2.bin_ena,
            (t2.val_produt * t1.val_vaz_natr).label('val_ena')
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(t2.bin_ena == 1). \
            join(t2). \
            order_by('dat_medicao','num_posto')
        
        df = pd.DataFrame(pd.read_sql(sql=stmt.statement, con=session.bind))
        
        df_ena = df.groupby(['dat_medicao','num_ssis','num_bacia','data_prev','modelo']).agg(dict(val_ena='sum')).reset_index()
        
        self.dados_ena_smap_hist = df_ena
    
    def query_enainter(self,data_inicial='2019-01-01',data_final='2020-01-23',ec45=0):
        
        session = Session(bind=engine)
        
        if ec45 == 0:
            t1 = aliased(EnaInter, name='t1')
        else:
            t1 = aliased(EnaInterEc45, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.num_ssis,
            t1.semana_op,
            t1.dias,
            t1.rev,
            t1.val_ena,
            t1.modelo,
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_enainter = pd.DataFrame(df)
        
    def query_pldinter(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)
        
        t1 = aliased(PldInter, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.modelo,
            t1.nom_deck,
            t1.discretizacao,
            t1.revisao,
            t1.pld_se,
            t1.pld_s,
            t1.pld_ne,
            t1.pld_n
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_pldinter = pd.DataFrame(df)
        
    def query_chuvaobs(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        '''
        Query de dados do Merge entre os dias indicados
        
        :data_inicial: Date, data inicial da consulta
        :data_final  : Date, data final da consulta
        '''
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaObs, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.nom_bacia,
            t1.hora,
            t1.val_prec,
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_chuva_obs = pd.DataFrame(df)
        
        return
        
    def query_chuvaimerg(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        '''
        Query de dados do IMERG entre os dias indicados
        
        :data_inicial: Date, data inicial da consulta
        :data_final  : Date, data final da consulta
        '''
        
        session = Session(bind=engine)
        
        t1 = aliased(ChuvaObsImerg, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.nom_bacia,
            t1.hora,
            t1.minu,
            t1.val_prec,
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_chuva_obs_imerg = pd.DataFrame(df)
        
        return
        
    def query_chuvagefs(self,data_inicial='2019-01-01',data_final='2020-01-23'):
        
        session = Session(bind=engine)

        t1 = aliased(ChuvaGefs, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.dat_prev,
            t1.rodada,
            t1.nom_bacia,
            t1.val_prec,
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_chuva_gefs = pd.DataFrame(df)
        
    def query_carga_sem(self, data_inicial='2019-01-01',data_final='2021-02-28'):
        
        session = Session(bind=engine)
        
        t1 = aliased(CargaSem, name='t1')
        
        stmt = session.query(
            t1.data_gera,
            t1.revisao,
            t1.num_semana,
            t1.num_ssis,
            t1.carga_med,
        ). \
            filter(t1.data_gera >= data_inicial). \
            filter(t1.data_gera <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_carga_sem = pd.DataFrame(df)
        
    def query_vazao_sem(self, data_inicial='2019-01-01',data_final='2021-02-28'):
        
        session = Session(bind=engine)
        
        t1 = aliased(VazaoSem, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,     # Data de referencia
            t1.data_ini,        # Data inicio da semana operativa
            t1.data_fim,        # Data fim da semana operativa
            t1.revisao,         # Numero da revisao do mes
            t1.ena_sem_se,      # ENA media semanal do SE
            t1.ena_sem_s,       # ENA media semanal do S
            t1.ena_sem_ne,      # ENA media semanal do NE
            t1.ena_sem_n,       # ENA media semanal do N
            t1.ena_sem_se_perc, # ENA media semanal percentual SE
            t1.ena_sem_s_perc,  # ENA media semanal percentual S
            t1.ena_sem_ne_perc, # ENA media semanal percentual NE
            t1.ena_sem_n_perc,  # ENA media semanal percentual N
            t1.ena_mes_se,      # ENA media mensal do SE
            t1.ena_mes_s,       # ENA media mensal do S
            t1.ena_mes_ne,      # ENA media mensal do NE
            t1.ena_mes_n,       # ENA media mensal do N
            t1.ena_mes_se_perc, # ENA media mensal percentual SE
            t1.ena_mes_s_perc,  # ENA media mensal percentual S
            t1.ena_mes_ne_perc, # ENA media mensal percentual NE
            t1.ena_mes_n_perc,  # ENA media mensal percentual N
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_vazao_sem = pd.DataFrame(df)
        
    def query_carga_mes(self, data_inicial='2019-01-01',data_final='2022-06-28'):
        
        session = Session(bind=engine)
        
        t1 = aliased(CargaMes, name='t1')
        
        stmt = session.query(
            t1.data_gera,
            t1.num_ssis,
            t1.mes,
            t1.ano,
            t1.carga_med,
        ). \
            filter(t1.data_gera >= data_inicial). \
            filter(t1.data_gera <= data_final)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_carga_mes = pd.DataFrame(df)
        
        return
        
    def query_previvaz(self, data_inicial='2019-01-01',data_final='2022-06-28',ano_inicial=2001):
        
        session = Session(bind=engine)
        
        t1 = aliased(HistPvv, name='t1')
        
        stmt = session.query(
            t1.dat_medicao,
            t1.num_posto,
            t1.num_semana,
            t1.num_ano,
            t1.val_vaz,
        ). \
            filter(t1.dat_medicao >= data_inicial). \
            filter(t1.dat_medicao <= data_final). \
            filter(t1.num_ano >= ano_inicial)

        df = pd.read_sql(sql=stmt.statement, con=session.bind)
        
        self.dados_pvv = pd.DataFrame(df)
        
        return